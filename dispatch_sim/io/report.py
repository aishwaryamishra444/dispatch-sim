"""Report writer — one .xlsx per scenario (per the team plan) plus a
comparison workbook. Blocks sheet holds the 96 simulated rows (data);
summary totals are live SUM formulas over the blocks sheet, so the workbook
recalculates if an analyst edits a block."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dispatch_sim.core.ledger import ScenarioResult

ARIAL = Font(name="Arial", size=10)
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="16202E")
NOTE = Font(name="Arial", size=9, italic=True, color="5B6B7F")
INR = '₹ #,##0;(₹ #,##0);"-"'
MWH = "0.000"

COLS = [  # (header, LedgerRow attr, number format)
    ("time", "time", None),
    ("scheduled_mwh", "scheduled_mwh", MWH),
    ("actual_gen_mwh", "actual_gen_mwh", MWH),
    ("delivered_mwh", "delivered_mwh", MWH),
    ("deviation_mwh", "deviation_mwh", "0.000;-0.000"),
    ("deviation_pct_of_avc", "deviation_pct_of_avc", "0.0"),
    ("ppa_revenue", "ppa_revenue", INR),
    ("dsm_receivable", "dsm_receivable", INR),
    ("dsm_payable", "dsm_payable", INR),
    ("dsm_penalty", "dsm_penalty", INR),
    ("soc_mwh", "soc_mwh", MWH),
    ("degradation", "degradation", INR),
    ("om", "om", INR),
    ("profit", "profit", INR),
]
SUM_COLS = ["scheduled_mwh", "actual_gen_mwh", "delivered_mwh", "ppa_revenue",
            "dsm_receivable", "dsm_payable", "dsm_penalty", "degradation",
            "om", "profit"]


def write_scenario_xlsx(result: ScenarioResult, path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "blocks"
    ws.freeze_panes = "A2"

    for j, (header, _, _) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=j, value=header)
        c.font, c.fill = HEAD, HEAD_FILL
        ws.column_dimensions[get_column_letter(j)].width = max(12, len(header) + 2)

    for i, row in enumerate(result.rows, start=2):
        for j, (_, attr, fmt) in enumerate(COLS, start=1):
            v = getattr(row, attr)
            c = ws.cell(row=i, column=j, value=v)
            c.font = ARIAL
            if fmt:
                c.number_format = fmt

    # ---- summary sheet: live formulas over blocks ----
    sm = wb.create_sheet("summary")
    sm.column_dimensions["A"].width = 26
    sm.column_dimensions["B"].width = 18
    sm["A1"], sm["A1"].font = result.name, Font(name="Arial", size=12, bold=True)
    r = 3
    for name in SUM_COLS:
        j = next(k for k, (h, _, _) in enumerate(COLS, start=1) if h == name)
        col = get_column_letter(j)
        sm.cell(row=r, column=1, value=f"total_{name}").font = ARIAL
        c = sm.cell(row=r, column=2, value=f"=SUM(blocks!{col}2:{col}97)")
        c.font = ARIAL
        c.number_format = INR if "mwh" not in name else MWH
        r += 1
    sm.cell(row=r + 1, column=1,
            value="Config snapshot (audit): " + str(result.config_snapshot)).font = NOTE
    sm.cell(row=r + 2, column=1,
            value="Data rows are simulated results; totals are live SUM formulas.").font = NOTE
    wb.save(path)


def write_comparison_xlsx(results: list[ScenarioResult], path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "comparison"
    headers = ["scenario", "gross_ppa", "dsm_receivable", "dsm_payable",
               "degradation", "om", "profit_per_day", "profit_per_year_330d"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill = HEAD, HEAD_FILL
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(h) + 2)

    for i, res in enumerate(results, start=2):
        vals = [res.name, res.total("ppa_revenue"), res.total("dsm_receivable"),
                res.total("dsm_payable"), res.total("degradation"), res.total("om")]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = ARIAL
            if j > 1:
                c.number_format = INR
        # profit as a live formula from this row's components
        c = ws.cell(row=i, column=7, value=f"=B{i}+C{i}-D{i}-E{i}-F{i}")
        c.font, c.number_format = ARIAL, INR
        c = ws.cell(row=i, column=8, value=f"=G{i}*330")
        c.font, c.number_format = ARIAL, INR

    n = len(results) + 2
    if len(results) >= 2 and results[1].total("profit") >= results[0].total("profit"):
        ws.cell(row=n + 1, column=1, value="CHECK: S2 did NOT come out worse than S1 "
                "on this day (expected on low-forecast-error days — review, per plan).").font = NOTE
    ws.cell(row=n + 2, column=1, value="Component values are simulated results; "
            "profit columns are live formulas.").font = NOTE
    wb.save(path)
